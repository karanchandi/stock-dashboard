"""
Global Value Screener v2
Single-pass yfinance pull (fundamentals + analyst data together),
then EDGAR insider data in a separate pass.
"""
import yfinance as yf
import pandas as pd
import time
import sys
import warnings
from ticker_universe import build_universe

warnings.filterwarnings("ignore")

# --- Rate limiting config ---
DELAY_PER_TICKER = 1.2      # seconds between yfinance calls (safe for 345 tickers)
DELAY_EVERY_50 = 10          # extra pause every 50 tickers
EDGAR_DELAY = 0.15           # seconds between EDGAR calls (SEC allows 10/sec)
MAX_RETRIES = 2              # retry failed tickers once

WEIGHTS = {
    "default": {
        "pe_score": 0.20, "pb_score": 0.20, "ev_ebitda_score": 0.15,
        "div_yield_score": 0.20, "debt_equity_score": 0.10, "52w_score": 0.15,
    },
    "REIT": {
        "pe_score": 0.10, "pb_score": 0.15, "ev_ebitda_score": 0.10,
        "div_yield_score": 0.30, "debt_equity_score": 0.15, "52w_score": 0.10,
        "ffo_score": 0.10,
    },
}


def safe_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return f if f == f else None
    except (ValueError, TypeError):
        return None


def fetch_all_data(ticker_info):
    """Single yfinance call: pull fundamentals + analyst data together."""
    ticker = ticker_info["ticker"]
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        if not info or info.get("regularMarketPrice") is None:
            return None

        price = safe_float(info.get("regularMarketPrice")) or safe_float(info.get("currentPrice"))
        if not price:
            return None

        high_52 = safe_float(info.get("fiftyTwoWeekHigh"))
        low_52 = safe_float(info.get("fiftyTwoWeekLow"))
        pct_from_52w_high = ((price - high_52) / high_52 * 100) if high_52 else None

        # Analyst upside calculation
        target_mean = safe_float(info.get("targetMeanPrice"))
        upside = round((target_mean - price) / price * 100, 2) if target_mean and price else None

        # Earnings estimate revisions (separate call but same ticker object)
        cq_growth = None
        ny_growth = None
        try:
            est = stock.earnings_estimate
            if est is not None and not est.empty:
                for idx, row in est.iterrows():
                    if str(idx) == "0q":
                        cq_growth = safe_float(row.get("growth"))
                    if str(idx) == "+1y":
                        ny_growth = safe_float(row.get("growth"))
        except Exception:
            pass

        # Recommendation trend shift
        buy_shift = None
        try:
            recs = stock.recommendations
            if recs is not None and len(recs) >= 2:
                cur = recs.iloc[0]
                pri = recs.iloc[1]
                cur_bulls = int(cur.get("strongBuy", 0)) + int(cur.get("buy", 0))
                pri_bulls = int(pri.get("strongBuy", 0)) + int(pri.get("buy", 0))
                buy_shift = cur_bulls - pri_bulls
        except Exception:
            pass

        # Earnings data
        next_earnings_date = None
        eps_estimate = None
        eps_actual = None
        eps_surprise_pct = None
        try:
            cal = stock.calendar
            if cal and "Earnings Date" in cal:
                dates = cal["Earnings Date"]
                if dates and len(dates) > 0:
                    next_earnings_date = str(dates[0])
            if cal and "Earnings Average" in cal:
                eps_estimate = safe_float(cal["Earnings Average"])
        except Exception:
            pass

        try:
            ed = stock.earnings_dates
            if ed is not None and len(ed) >= 2:
                # Most recent completed quarter (second row, first is upcoming)
                recent = ed.iloc[1]
                if recent.get("Reported EPS") is not None:
                    eps_actual = safe_float(recent.get("Reported EPS"))
                    eps_surprise_pct = safe_float(recent.get("Surprise(%)"))
        except Exception:
            pass

        return {
            # Fundamentals
            "ticker": ticker,
            "name": info.get("shortName", ""),
            "exchange": ticker_info["exchange"],
            "sector": ticker_info["sector"],
            "subsector": ticker_info["subsector"],
            "currency": info.get("currency", ""),
            "price": price,
            "market_cap": safe_float(info.get("marketCap")),
            "pe_ratio": safe_float(info.get("trailingPE")),
            "forward_pe": safe_float(info.get("forwardPE")),
            "pb_ratio": safe_float(info.get("priceToBook")),
            "ev_ebitda": safe_float(info.get("enterpriseToEbitda")),
            "dividend_yield": safe_float(info.get("dividendYield")),
            "dividend_rate": safe_float(info.get("dividendRate")),
            "payout_ratio": safe_float(info.get("payoutRatio")),
            "debt_to_equity": safe_float(info.get("debtToEquity")),
            "return_on_equity": safe_float(info.get("returnOnEquity")),
            "revenue_growth": safe_float(info.get("revenueGrowth")),
            "earnings_growth": safe_float(info.get("earningsGrowth")),
            "profit_margin": safe_float(info.get("profitMargins")),
            "operating_margin": safe_float(info.get("operatingMargins")),
            "free_cashflow": safe_float(info.get("freeCashflow")),
            "total_revenue": safe_float(info.get("totalRevenue")),
            "net_income": safe_float(info.get("netIncomeToCommon")),
            "52w_high": high_52,
            "52w_low": low_52,
            "pct_from_52w_high": pct_from_52w_high,
            "avg_volume": safe_float(info.get("averageVolume")),
            "beta": safe_float(info.get("beta")),
            # Analyst data (pulled from same .info + earnings_estimate)
            "target_price": target_mean,
            "target_high": safe_float(info.get("targetHighPrice")),
            "target_low": safe_float(info.get("targetLowPrice")),
            "upside_pct": upside,
            "analyst_rating": safe_float(info.get("recommendationMean")),
            "analyst_consensus": info.get("recommendationKey", ""),
            "num_analysts": info.get("numberOfAnalystOpinions"),
            "forward_eps": safe_float(info.get("forwardEps")),
            "cq_earnings_growth": cq_growth,
            "next_yr_earnings_growth": ny_growth,
            "buy_shift_mom": buy_shift,
            # Earnings data
            "next_earnings_date": next_earnings_date,
            "eps_estimate": eps_estimate,
            "eps_actual": eps_actual,
            "eps_surprise_pct": eps_surprise_pct,
            "trailing_eps": safe_float(info.get("trailingEps")),
            "eps_current_year": safe_float(info.get("epsCurrentYear")),
        }
    except Exception:
        return None


def score_value(row):
    subsector = row.get("subsector", "")
    weights = WEIGHTS.get(subsector, WEIGHTS["default"])
    scores = {}

    pe = safe_float(row.get("pe_ratio"))
    if pe is not None and 0 < pe < 100:
        scores["pe_score"] = max(0, min(100, 100 - (pe / 30 * 100)))
    else:
        scores["pe_score"] = None

    pb = safe_float(row.get("pb_ratio"))
    if pb is not None and 0 < pb < 50:
        scores["pb_score"] = max(0, min(100, 100 - (pb / 5 * 100)))
    else:
        scores["pb_score"] = None

    ev = safe_float(row.get("ev_ebitda"))
    if ev is not None and 0 < ev < 100:
        scores["ev_ebitda_score"] = max(0, min(100, 100 - (ev / 20 * 100)))
    else:
        scores["ev_ebitda_score"] = None

    dy = safe_float(row.get("dividend_yield"))
    if dy is not None and dy > 0:
        # yfinance returns dividendYield as percentage (e.g. 5.22 = 5.22%)
        # Score: 10% yield = 100 score
        scores["div_yield_score"] = min(100, dy / 10 * 100)
    else:
        scores["div_yield_score"] = 0

    de = safe_float(row.get("debt_to_equity"))
    if de is not None and de >= 0:
        cap = 300 if subsector == "REIT" else 200
        scores["debt_equity_score"] = max(0, min(100, 100 - (de / cap * 100)))
    else:
        scores["debt_equity_score"] = None

    pct = safe_float(row.get("pct_from_52w_high"))
    if pct is not None:
        scores["52w_score"] = min(100, max(0, abs(pct) * 2))
    else:
        scores["52w_score"] = None

    if subsector == "REIT":
        fcf = safe_float(row.get("free_cashflow"))
        mcap = safe_float(row.get("market_cap"))
        if fcf and mcap and mcap > 0:
            ffo_yield = fcf / mcap * 100
            scores["ffo_score"] = min(100, max(0, ffo_yield * 10))
        else:
            scores["ffo_score"] = None

    total_weight = 0
    total_score = 0
    for key, weight in weights.items():
        if scores.get(key) is not None:
            total_score += scores[key] * weight
            total_weight += weight

    composite = (total_score / total_weight) if total_weight > 0 else None
    return composite, scores


def score_analyst(row):
    """Score analyst sentiment 0-100."""
    weights = {"upside": 0.30, "rating": 0.25, "growth": 0.25, "momentum": 0.20}
    scores = {}

    upside = safe_float(row.get("upside_pct"))
    if upside is not None:
        scores["upside"] = min(100, max(0, upside * 2))

    rating = safe_float(row.get("analyst_rating"))
    if rating is not None and 1 <= rating <= 5:
        scores["rating"] = max(0, (5 - rating) / 4 * 100)

    eg = safe_float(row.get("next_yr_earnings_growth")) or safe_float(row.get("earnings_growth"))
    if eg is not None:
        scores["growth"] = min(100, max(0, 50 + eg * 200))

    shift = row.get("buy_shift_mom")
    if shift is not None and pd.notna(shift):
        scores["momentum"] = min(100, max(0, 50 + float(shift) * 15))

    tw = sum(weights[k] for k in scores)
    if tw == 0:
        return None
    return round(sum(scores[k] * weights[k] for k in scores) / tw, 2)


def categorize_market_cap(mc):
    if mc is None: return "Unknown"
    if mc >= 10e9: return "Large Cap"
    if mc >= 2e9: return "Mid Cap"
    if mc >= 300e6: return "Small Cap"
    return "Micro Cap"


def run_screener():
    universe = build_universe()
    total = len(universe)
    est_minutes = round(total * DELAY_PER_TICKER / 60 + total // 50 * DELAY_EVERY_50 / 60, 1)
    print(f"Screening {total} tickers (fundamentals + analyst in single pass)")
    print(f"Estimated time: ~{est_minutes} minutes with safe rate limiting\n")

    results = []
    errors = []

    for i, ticker_info in enumerate(universe):
        ticker = ticker_info["ticker"]
        pct = (i + 1) / total * 100
        sys.stdout.write(f"\r[{pct:5.1f}%] {ticker:<15} ({i+1}/{total})")
        sys.stdout.flush()

        data = fetch_all_data(ticker_info)
        if data:
            composite, component_scores = score_value(data)
            data["value_score"] = round(composite, 2) if composite else None
            for k, v in component_scores.items():
                data[k] = round(v, 2) if v is not None else None
            data["market_cap_tier"] = categorize_market_cap(data.get("market_cap"))
            data["analyst_score"] = score_analyst(data)
            results.append(data)
        else:
            errors.append(ticker_info)

        # Rate limiting
        time.sleep(DELAY_PER_TICKER)
        if (i + 1) % 50 == 0:
            print(f"\n  ... pausing {DELAY_EVERY_50}s to avoid rate limit (at {i+1}/{total}) ...")
            time.sleep(DELAY_EVERY_50)

    # Retry failed tickers once with longer delay
    if errors and MAX_RETRIES > 0:
        print(f"\n\nRetrying {len(errors)} failed tickers with longer delay...")
        time.sleep(30)
        still_failed = []
        for j, ticker_info in enumerate(errors):
            ticker = ticker_info["ticker"]
            sys.stdout.write(f"\r  Retry: {ticker:<15} ({j+1}/{len(errors)})")
            sys.stdout.flush()
            data = fetch_all_data(ticker_info)
            if data:
                composite, component_scores = score_value(data)
                data["value_score"] = round(composite, 2) if composite else None
                for k, v in component_scores.items():
                    data[k] = round(v, 2) if v is not None else None
                data["market_cap_tier"] = categorize_market_cap(data.get("market_cap"))
                data["analyst_score"] = score_analyst(data)
                results.append(data)
            else:
                still_failed.append(ticker)
            time.sleep(2.0)
        errors = still_failed

    print(f"\n\nCompleted: {len(results)} successful, {len(errors)} failed")
    if errors:
        print(f"Failed ({len(errors)}): {', '.join(str(e) for e in errors[:30])}")

    df = pd.DataFrame(results)
    if df.empty:
        print("No data retrieved.")
        return df

    if "dividend_yield" in df.columns:
        df["dividend_yield_pct"] = df["dividend_yield"].apply(
            lambda x: round(x * 100, 2) if pd.notna(x) else None
        )

    # --- EDGAR Insider Data (US tickers only) ---
    print("\n" + "=" * 50)
    print("EDGAR Insider Scan (US tickers only)")
    print("=" * 50)
    from edgar_insider import run_insider_scan
    tickers_list = df["ticker"].tolist()
    insider_data = run_insider_scan(tickers_list)
    insider_cols = ["buy_count", "sell_count", "unique_buyers", "total_buy_value",
                    "total_sell_value", "net_value", "insider_score"]
    for col in insider_cols:
        df[col] = df["ticker"].map(lambda t, c=col: insider_data.get(t, {}).get(c))

    # --- Combined Score ---
    def combined_score(row):
        v, a, i = row.get("value_score"), row.get("analyst_score"), row.get("insider_score")
        parts, weights = [], []
        if pd.notna(v): parts.append(v); weights.append(0.50)
        if pd.notna(a): parts.append(a); weights.append(0.30)
        if pd.notna(i): parts.append(i); weights.append(0.20)
        if not parts: return None
        tw = sum(weights)
        return round(sum(p * w for p, w in zip(parts, weights)) / tw, 2)

    df["combined_score"] = df.apply(combined_score, axis=1)
    df = df.sort_values("combined_score", ascending=False, na_position="last")
    return df


def save_to_excel(df, filename="global_value_screener.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    display_cols = [
        "ticker", "name", "exchange", "subsector", "market_cap_tier", "currency",
        "price", "market_cap",
        "combined_score", "value_score", "analyst_score", "insider_score",
        "pe_ratio", "forward_pe", "pb_ratio", "ev_ebitda",
        "dividend_yield_pct", "payout_ratio", "debt_to_equity", "return_on_equity",
        "profit_margin", "52w_high", "52w_low", "pct_from_52w_high",
        "target_price", "upside_pct", "analyst_consensus", "num_analysts",
        "earnings_growth", "revenue_growth",
        "buy_count", "sell_count", "net_value",
        "beta", "avg_volume",
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    data_font = Font(name="Arial", size=10)
    top_fill = PatternFill("solid", fgColor="C6EFCE")
    border = Border(bottom=Side(style="thin", color="D9D9D9"))

    col_headers = {
        "ticker": "Ticker", "name": "Company", "exchange": "Exchange",
        "subsector": "Subsector", "market_cap_tier": "Cap Tier", "currency": "Ccy",
        "price": "Price", "market_cap": "Market Cap",
        "combined_score": "Combined Score", "value_score": "Value Score",
        "analyst_score": "Analyst Score", "insider_score": "Insider Score",
        "pe_ratio": "P/E", "forward_pe": "Fwd P/E", "pb_ratio": "P/B",
        "ev_ebitda": "EV/EBITDA", "dividend_yield_pct": "Div Yield %",
        "payout_ratio": "Payout Ratio", "debt_to_equity": "D/E",
        "return_on_equity": "ROE", "profit_margin": "Margin",
        "52w_high": "52W High", "52w_low": "52W Low",
        "pct_from_52w_high": "% From 52W High",
        "target_price": "Target Price", "upside_pct": "Upside %",
        "analyst_consensus": "Consensus", "num_analysts": "# Analysts",
        "earnings_growth": "Earnings Growth", "revenue_growth": "Rev Growth",
        "buy_count": "Insider Buys", "sell_count": "Insider Sells",
        "net_value": "Insider Net $",
        "beta": "Beta", "avg_volume": "Avg Vol",
    }

    tabs = {"All": df}
    for sub in sorted(df["subsector"].dropna().unique()):
        tabs[sub] = df[df["subsector"] == sub]

    first = True
    for tab_name, tab_df in tabs.items():
        if tab_df.empty:
            continue
        ws = wb.active if first else wb.create_sheet()
        ws.title = tab_name[:31]
        first = False

        available_cols = [c for c in display_cols if c in tab_df.columns]
        headers = [col_headers.get(c, c) for c in available_cols]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, (_, row) in enumerate(tab_df.iterrows(), 2):
            for col_idx, col_name in enumerate(available_cols, 1):
                val = row.get(col_name)
                if pd.isna(val):
                    val = None
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = border
                if row_idx <= 11:
                    cell.fill = top_fill

        for col_idx in range(1, len(available_cols) + 1):
            max_len = max(len(str(headers[col_idx - 1])), 12)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        ws.auto_filter.ref = f"A1:{get_column_letter(len(available_cols))}{len(tab_df)+1}"
        ws.freeze_panes = "A2"

    wb.save(filename)
    print(f"\nSaved to {filename}")


if __name__ == "__main__":
    df = run_screener()
    if not df.empty:
        output = "/home/claude/stock_screener/global_value_screener.xlsx"
        save_to_excel(df, output)
        csv = "/home/claude/stock_screener/global_value_screener.csv"
        df.to_csv(csv, index=False)
        print(f"CSV saved to {csv}")
        print(f"\nTop 25 by Combined Score:")
        cols = ["ticker", "name", "subsector", "combined_score", "value_score",
                "analyst_score", "insider_score", "upside_pct", "price"]
        cols = [c for c in cols if c in df.columns]
        print(df.head(25)[cols].to_string())
